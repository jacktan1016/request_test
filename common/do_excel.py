# -*- coding:utf-8 -*-
"""
@author: JakeTan
@file: do_excel.py
@time: 2020/2/5 14:32
@desc: 封装excel的表格
"""
import openpyxl
from common import logger

logger = logger.get_logger('excel')

class Case:
    """封装测试用例"""
    def __init__(self):
        self.id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None

class Do_excel:

    def __init__(self,file_name):
        self.file_name = file_name
        try:
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            logger.info("没有找到文件:{}".format(file_name))
            raise e

    def read_excel(self,sheet_name):
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        cases = []
        for i in range(2, max_row + 1):
            case = Case()
            case.id = sheet.cell(row=i, column=1).value
            case.title = sheet.cell(row=i, column=2).value
            case.url = sheet.cell(row=i, column=3).value
            case.data = sheet.cell(row=i, column=4).value
            case.method = sheet.cell(row=i, column=5).value
            case.expected = sheet.cell(row=i, column=6).value
            if type(case.expected) == int:
                case.expected = str(case.expected)
            cases.append(case)
        return cases

    def write_excel(self,sheet_name,row,result,testresult):
        sheet = self.workbook[sheet_name]
        sheet.cell(row=row,column = 7).value = result
        sheet.cell(row=row,column=8).value = testresult
        self.workbook.save(filename=self.file_name)


if __name__=="__main__":
    from common import contants
    from common.requests_pack import DoRequest
    import json,os
    request = DoRequest()
    excel = Do_excel(contants.case_path)
    login_cases = excel.read_excel("login")
    for login_case in login_cases:
        # print(login_case.id,login_case.method,login_case.expected)
        print("第{}条测试用例".format(login_case.id))
        print(login_case.method,login_case.url,login_case.data)
        res = request.request_p(login_case.method,login_case.url,login_case.data)
        if res.text ==login_case.expected:
            excel.write_excel("login",login_case.id+1,res.text,"pass")
            print("第{}条测试用例:通过".format(login_case.id))
        else:
            excel.write_excel("login", login_case.id+1, res.text, "fail")
            print("第{}条测试用例:失败".format(login_case.id))









